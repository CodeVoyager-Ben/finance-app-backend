from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Sum, Q, F
from django.db.models.functions import TruncMonth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from decimal import Decimal

from apps.transactions.models import Account, Category, Transaction
from apps.investments.models import InvestmentAccount, InvestmentHolding
from apps.investments.services import to_cny
from apps.lending.models import LendingRecord


class BalanceSheetView(APIView):
    """个人资产负债表"""

    def get(self, request):
        user = request.user
        target_date = request.query_params.get('date')

        # Parse target date
        if target_date:
            try:
                target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
            except ValueError:
                target_date = None
        if not target_date:
            target_date = date.today()

        # --- 资产：现金及银行 ---
        accounts = Account.objects.filter(user=user, is_active=True, exclude_from_reports=False)
        cash_assets = []
        total_cash = 0

        for acc in accounts:
            balance = acc.balance
            # 反向计算：减去目标日期之后的交易净变动
            if target_date < date.today():
                future_tx = Transaction.objects.filter(
                    account=acc, date__gt=target_date
                ).aggregate(
                    income=Sum('amount', filter=Q(transaction_type='income')),
                    expense=Sum('amount', filter=Q(transaction_type='expense')),
                )
                net_change = (future_tx['income'] or 0) - (future_tx['expense'] or 0)
                balance = balance - net_change

            if balance > 0:
                cash_assets.append({
                    'name': acc.name,
                    'type': acc.get_account_type_display(),
                    'amount': float(balance),
                })
                total_cash += float(balance)

        # --- 资产：投资持仓 ---
        holdings = InvestmentHolding.objects.filter(
            investment_account__user=user
        ).select_related('investment_account', 'investment_account__asset_type')
        invest_assets = []
        total_invest = 0
        by_asset_category = {}

        for h in holdings:
            mv_cny = float(to_cny(h.market_value, h.effective_currency))
            if mv_cny > 0:
                at = h.investment_account.asset_type
                type_name = at.name if at else '其他'
                invest_assets.append({
                    'name': f'{h.name}({h.symbol})',
                    'type': type_name,
                    'currency': h.effective_currency,
                    'market_value': float(h.market_value),
                    'market_value_cny': mv_cny,
                    'cost_value': float(h.cost_value),
                    'profit_loss': float(h.profit_loss),
                })
                total_invest += mv_cny

                # 按资产大类分组
                cat = at.category if at else 'other'
                cat_display = dict(InvestmentAccount.asset_type.field.related_model.CATEGORY_CHOICES).get(cat, '其他') if at else '其他'
                if cat not in by_asset_category:
                    by_asset_category[cat] = {'category': cat_display, 'amount': 0}
                by_asset_category[cat]['amount'] += mv_cny

        # 投资账户余额
        invest_balance = 0
        for a in InvestmentAccount.objects.filter(user=user, is_active=True).select_related('asset_type'):
            invest_balance += float(to_cny(a.balance, a.currency))

        # --- 资产：应收款项（借出未收回，按对方合并）---
        lend_records = LendingRecord.objects.filter(
            user=user, record_type='lend', status__in=['outstanding', 'partial']
        )
        receivables_map = {}
        for r in lend_records:
            remaining = float(r.remaining_amount)
            if remaining > 0:
                key = r.counterparty
                if key not in receivables_map:
                    receivables_map[key] = {'name': key, 'amount': 0, 'count': 0}
                receivables_map[key]['amount'] += remaining
                receivables_map[key]['count'] += 1
        receivables = list(receivables_map.values())
        total_receivable = sum(item['amount'] for item in receivables)

        total_assets = total_cash + total_invest + invest_balance + total_receivable

        # --- 资产配置占比 ---
        allocation = []
        if total_assets > 0:
            allocation = [
                {'category': '现金及银行', 'amount': total_cash,
                 'percentage': round(total_cash / total_assets * 100, 1)},
                {'category': '证券账户余额', 'amount': invest_balance,
                 'percentage': round(invest_balance / total_assets * 100, 1)},
            ]
            for cat_data in by_asset_category.values():
                if cat_data['amount'] > 0:
                    allocation.append({
                        'category': cat_data['category'],
                        'amount': cat_data['amount'],
                        'percentage': round(cat_data['amount'] / total_assets * 100, 1),
                    })
            if total_receivable > 0:
                allocation.append({
                    'category': '应收款项',
                    'amount': total_receivable,
                    'percentage': round(total_receivable / total_assets * 100, 1),
                })
            allocation = [a for a in allocation if a['amount'] > 0]

        # --- 负债 (透支账户) ---
        liabilities = []
        total_liabilities = 0

        for acc in accounts:
            balance = acc.balance
            if target_date < date.today():
                future_tx = Transaction.objects.filter(
                    account=acc, date__gt=target_date
                ).aggregate(
                    income=Sum('amount', filter=Q(transaction_type='income')),
                    expense=Sum('amount', filter=Q(transaction_type='expense')),
                )
                net_change = (future_tx['income'] or 0) - (future_tx['expense'] or 0)
                balance = balance - net_change

            if balance < 0:
                liabilities.append({
                    'name': acc.name,
                    'type': acc.get_account_type_display(),
                    'amount': abs(float(balance)),
                })
                total_liabilities += abs(float(balance))

        # --- 负债：应付款项（借入未归还，按对方合并）---
        borrow_records = LendingRecord.objects.filter(
            user=user, record_type='borrow', status__in=['outstanding', 'partial']
        )
        payables_map = {}
        for r in borrow_records:
            remaining = float(r.remaining_amount)
            if remaining > 0:
                key = r.counterparty
                if key not in payables_map:
                    payables_map[key] = {'name': key, 'amount': 0, 'count': 0}
                payables_map[key]['amount'] += remaining
                payables_map[key]['count'] += 1
        payables = list(payables_map.values())
        total_payable = sum(item['amount'] for item in payables)

        total_liabilities += total_payable

        net_worth = total_assets - total_liabilities

        # --- 财务比率 ---
        ratios = self._calculate_ratios(user, total_assets, total_liabilities, total_cash, total_invest)

        # --- 净资产环比变化 ---
        net_worth_change = self._calculate_net_worth_change(user, net_worth)

        return Response({
            'date': str(target_date),
            'assets': {
                'cash': {'items': cash_assets, 'total': total_cash},
                'invest_balance': invest_balance,
                'investments': {'items': invest_assets, 'total': total_invest},
                'receivables': {'items': receivables, 'total': total_receivable},
                'total': total_assets,
                'allocation': allocation,
            },
            'liabilities': {
                'items': liabilities,
                'payables': {'items': payables, 'total': total_payable},
                'total': total_liabilities,
            },
            'net_worth': net_worth,
            'ratios': ratios,
            'net_worth_change': net_worth_change,
        })

    def _calculate_ratios(self, user, total_assets, total_liabilities, total_cash, total_invest):
        """计算财务健康比率"""
        # 负债率
        debt_ratio = round(total_liabilities / total_assets, 4) if total_assets > 0 else 0

        # 流动性比率 = 现金类资产 / 总负债
        current_ratio = round(total_cash / total_liabilities, 2) if total_liabilities > 0 else float('inf')

        # 储蓄率 = (月收入 - 月支出) / 月收入
        now = date.today()
        month_start = now.replace(day=1)
        month_tx = Transaction.objects.filter(
            user=user, date__gte=month_start, date__lte=now
        ).aggregate(
            income=Sum('amount', filter=Q(transaction_type='income')),
            expense=Sum('amount', filter=Q(transaction_type='expense')),
        )
        month_income = float(month_tx['income'] or 0)
        month_expense = float(month_tx['expense'] or 0)
        savings_ratio = round((month_income - month_expense) / month_income, 4) if month_income > 0 else 0

        # 投资占比
        investment_ratio = round(total_invest / total_assets, 4) if total_assets > 0 else 0

        # 综合健康评级
        health_level = '优秀'
        if debt_ratio > 0.5 or savings_ratio < 0:
            health_level = '危险'
        elif debt_ratio > 0.3 or savings_ratio < 0.1:
            health_level = '注意'
        elif debt_ratio < 0.1 and savings_ratio > 0.3:
            health_level = '优秀'
        else:
            health_level = '良好'

        return {
            'debt_ratio': debt_ratio,
            'current_ratio': current_ratio if current_ratio != float('inf') else 999,
            'savings_ratio': savings_ratio,
            'investment_ratio': investment_ratio,
            'month_income': month_income,
            'month_expense': month_expense,
            'health_level': health_level,
        }

    def _calculate_net_worth_change(self, user, current_net_worth):
        """计算净资产环比变化"""
        now = date.today()
        # 上月月末
        if now.month == 1:
            last_month_end = now.replace(year=now.year - 1, month=12, day=28)
        else:
            last_month_end = now.replace(month=now.month - 1, day=28)
        # 确保不超过该月实际天数
        import calendar
        last_day = calendar.monthrange(last_month_end.year, last_month_end.month)[1]
        last_month_end = last_month_end.replace(day=last_day)

        # 上月净资产 = 当前净资产 - 本月交易净变动
        month_tx = Transaction.objects.filter(
            user=user, date__gt=last_month_end
        ).aggregate(
            income=Sum('amount', filter=Q(transaction_type='income')),
            expense=Sum('amount', filter=Q(transaction_type='expense')),
        )
        month_net = float((month_tx['income'] or 0) - (month_tx['expense'] or 0))
        previous_net_worth = current_net_worth - month_net

        change = current_net_worth - previous_net_worth
        change_pct = round(change / abs(previous_net_worth) * 100, 2) if previous_net_worth != 0 else 0

        return {
            'current': current_net_worth,
            'previous': round(previous_net_worth, 2),
            'change': round(change, 2),
            'change_pct': change_pct,
        }


class NetWorthHistoryView(APIView):
    """净资产历史趋势"""

    def get(self, request):
        user = request.user
        months = int(request.query_params.get('months', 12))

        # 当前净资产
        accounts = Account.objects.filter(user=user, is_active=True, exclude_from_reports=False)
        invest_accounts = InvestmentAccount.objects.filter(user=user, is_active=True)
        holdings = InvestmentHolding.objects.filter(
            investment_account__user=user
        ).select_related('investment_account')

        current_cash = sum(float(a.balance) for a in accounts)
        current_invest_balance = sum(float(to_cny(a.balance, a.currency)) for a in invest_accounts)
        current_invest_holdings = sum(float(to_cny(h.market_value, h.effective_currency)) for h in holdings)
        current_assets = current_cash + current_invest_balance + current_invest_holdings

        # 负债：余额为负的账户
        current_liabilities = sum(abs(float(a.balance)) for a in accounts if a.balance < 0)
        # 资产中不包含负余额部分
        current_positive_cash = sum(float(a.balance) for a in accounts if a.balance > 0)
        current_total_assets = current_positive_cash + current_invest_balance + current_invest_holdings

        # 借贷数据
        from apps.lending.models import LendingRecord
        lend_remaining = sum(
            float(r.remaining_amount) for r in
            LendingRecord.objects.filter(user=user, record_type='lend', status__in=['outstanding', 'partial'])
        )
        borrow_remaining = sum(
            float(r.remaining_amount) for r in
            LendingRecord.objects.filter(user=user, record_type='borrow', status__in=['outstanding', 'partial'])
        )
        current_total_assets += lend_remaining
        current_liabilities += borrow_remaining

        # 获取过去N个月的所有交易，按月分组
        from datetime import timedelta
        start_date = date.today() - timedelta(days=months * 31)

        monthly_tx = Transaction.objects.filter(
            user=user, date__gte=start_date
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            income=Sum('amount', filter=Q(transaction_type='income')),
            expense=Sum('amount', filter=Q(transaction_type='expense')),
        ).order_by('month')

        # 构建月度交易净变动字典
        tx_by_month = {}
        for m in monthly_tx:
            month_key = m['month'].strftime('%Y-%m')
            tx_by_month[month_key] = float(m['income'] or 0) - float(m['expense'] or 0)

        # 生成月度列表
        import calendar
        history = []
        today = date.today()

        # 累积的未来交易变动（从当前月份往回减）
        cumulative_future = 0

        for i in range(months):
            # 从最远月份开始计算
            past = today - timedelta(days=(months - i) * 31)
            # 规范化到月初
            month_date = past.replace(day=1)
            month_key = month_date.strftime('%Y-%m')

            # 累加该月及之后所有的交易净变动
            net_change = 0
            for mk, nv in tx_by_month.items():
                if mk >= month_key:
                    net_change += nv

            assets_at_month = current_total_assets - (cumulative_future - net_change)
            liabilities_at_month = current_liabilities  # 简化：负债按当前值

            # 更精确的负债：也要扣除未来变动中影响负余额账户的部分
            # 简化处理：负债按比例调整
            if current_total_assets > 0:
                liability_ratio = current_liabilities / current_total_assets
                liabilities_at_month = max(0, assets_at_month * liability_ratio)
            else:
                liabilities_at_month = 0

            cumulative_future += tx_by_month.get(month_key, 0)

            net_worth_at_month = assets_at_month - liabilities_at_month

            history.append({
                'month': month_key,
                'assets': round(assets_at_month, 2),
                'liabilities': round(liabilities_at_month, 2),
                'net_worth': round(net_worth_at_month, 2),
            })

        # 添加当前月份
        history.append({
            'month': today.strftime('%Y-%m'),
            'assets': round(current_total_assets, 2),
            'liabilities': round(current_liabilities, 2),
            'net_worth': round(current_total_assets - current_liabilities, 2),
        })

        return Response({'history': history})


class ExportExcelView(APIView):
    """导出Excel报表"""

    def get(self, request):
        export_type = request.query_params.get('type', 'transactions')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if export_type == 'transactions':
            return self._export_transactions(request.user, start_date, end_date)
        elif export_type == 'balance_sheet':
            return self._export_balance_sheet(request.user)
        else:
            return Response({'error': '不支持的导出类型'}, status=400)

    def _export_transactions(self, user, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        ws.title = '收支明细'

        # Headers
        headers = ['日期', '类型', '分类', '账户', '金额', '备注']
        header_fill = PatternFill(start_color='1677FF', end_color='1677FF', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data
        transactions = Transaction.objects.filter(user=user).select_related('account', 'category')
        if start_date:
            transactions = transactions.filter(date__gte=start_date)
        if end_date:
            transactions = transactions.filter(date__lte=end_date)

        for row, t in enumerate(transactions, 2):
            data = [
                str(t.date),
                t.get_transaction_type_display(),
                t.category.name if t.category else '',
                t.account.name,
                float(t.amount),
                t.note,
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 5:  # amount column
                    if t.transaction_type == 'expense':
                        cell.font = Font(color='FF4D4F')
                    else:
                        cell.font = Font(color='52C41A')

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'transactions_{datetime.now().strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def _export_balance_sheet(self, user):
        wb = Workbook()
        ws = wb.active
        ws.title = '资产负债表'

        title_font = Font(bold=True, size=16)
        section_font = Font(bold=True, size=13, color='1677FF')
        section_font_red = Font(bold=True, size=13, color='FF4D4F')
        header_fill = PatternFill(start_color='F0F5FF', end_color='F0F5FF', fill_type='solid')
        header_fill_red = PatternFill(start_color='FFF1F0', end_color='FFF1F0', fill_type='solid')
        ratio_fill = PatternFill(start_color='F6F6F6', end_color='F6F6F6', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Title
        ws.merge_cells('A1:D1')
        ws.cell(row=1, column=1, value='个人资产负债表').font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=1, value=f'生成日期: {datetime.now().strftime("%Y-%m-%d")}')

        # === 资产部分 ===
        row = 4
        ws.cell(row=row, column=1, value='一、资产').font = section_font
        row += 1
        for col, header in enumerate(['项目', '类型', '金额', '备注'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.border = thin_border

        row += 1
        accounts = Account.objects.filter(user=user, is_active=True, exclude_from_reports=False)
        total_cash = 0
        for acc in accounts:
            if acc.balance > 0:
                for col, val in enumerate([acc.name, acc.get_account_type_display(), float(acc.balance), ''], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                total_cash += float(acc.balance)
                row += 1

        # Investment holdings
        holdings = InvestmentHolding.objects.filter(investment_account__user=user).select_related('investment_account', 'investment_account__asset_type')
        total_invest = 0
        for h in holdings:
            mv_cny = float(to_cny(h.market_value, h.effective_currency))
            if mv_cny > 0:
                at = h.investment_account.asset_type
                type_name = at.name if at else '其他'
                cur = h.effective_currency
                note_parts = [f'成本:{float(h.cost_value):.2f} 盈亏:{float(h.profit_loss):.2f}']
                if cur != 'CNY':
                    note_parts.append(f'原币:{float(h.market_value):.2f} {cur}')
                for col, val in enumerate([
                    f'{h.name}({h.symbol})',
                    type_name,
                    mv_cny,
                    ' '.join(note_parts),
                ], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                total_invest += mv_cny
                row += 1

        # Investment account balance
        invest_balance = sum(
            float(to_cny(a.balance, a.currency))
            for a in InvestmentAccount.objects.filter(user=user, is_active=True).select_related('asset_type')
        )

        # Receivables (lend outstanding)
        total_receivable = 0
        for r in LendingRecord.objects.filter(user=user, record_type='lend', status__in=['outstanding', 'partial']):
            remaining = float(r.remaining_amount)
            if remaining > 0:
                for col, val in enumerate([f'应收-{r.counterparty}', '借出', remaining, r.reason or ''], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                total_receivable += remaining
                row += 1

        total_assets = total_cash + total_invest + invest_balance + total_receivable

        # Total assets row
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws.cell(row=row, column=1, value='资产合计')
        ws.cell(row=row, column=3, value=total_assets)
        row += 1

        # === 负债部分 ===
        row += 1
        ws.cell(row=row, column=1, value='二、负债').font = section_font_red
        row += 1
        for col, header in enumerate(['项目', '类型', '金额', '备注'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill_red
            cell.font = Font(bold=True)
            cell.border = thin_border

        row += 1
        total_liabilities = 0
        for acc in accounts:
            if acc.balance < 0:
                amount = abs(float(acc.balance))
                for col, val in enumerate([acc.name, acc.get_account_type_display(), amount, ''], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    if col == 3:
                        cell.font = Font(color='FF4D4F')
                total_liabilities += amount
                row += 1

        # Payables (borrow outstanding)
        for r in LendingRecord.objects.filter(user=user, record_type='borrow', status__in=['outstanding', 'partial']):
            remaining = float(r.remaining_amount)
            if remaining > 0:
                for col, val in enumerate([f'应付-{r.counterparty}', '借入', remaining, r.reason or ''], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    if col == 3:
                        cell.font = Font(color='FF4D4F')
                total_liabilities += remaining
                row += 1

        # Total liabilities row
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = Font(bold=True, color='FF4D4F')
            cell.fill = header_fill_red
        ws.cell(row=row, column=1, value='负债合计')
        ws.cell(row=row, column=3, value=total_liabilities)
        row += 1

        # === 净资产 ===
        row += 1
        net_worth = total_assets - total_liabilities
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = Font(bold=True, size=14, color='1677FF')
        ws.cell(row=row, column=1, value='净资产（资产 - 负债）')
        ws.cell(row=row, column=3, value=net_worth)
        row += 1

        # === 财务比率摘要 ===
        row += 1
        ws.cell(row=row, column=1, value='三、财务健康指标').font = Font(bold=True, size=13, color='722ED1')
        row += 1
        ratio_items = [
            ('资产负债率', f'{round(total_liabilities/total_assets*100, 1)}%' if total_assets > 0 else 'N/A',
             '<30%健康, 30-50%注意, >50%危险'),
            ('流动性比率', f'{round(total_cash/total_liabilities, 2)}' if total_liabilities > 0 else '无负债',
             '>2健康, 1-2一般, <1警告'),
            ('投资占比', f'{round(total_invest/total_assets*100, 1)}%' if total_assets > 0 else 'N/A',
             '投资资产占总资产比例'),
        ]
        for name, value, note in ratio_items:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = ratio_fill
                ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=3, value=value)
            ws.cell(row=row, column=4, value=note)
            row += 1

        # Column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 24

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'balance_sheet_{datetime.now().strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
